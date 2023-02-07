import pandas as pd
import requests
import json
from finnomena_api import finnomenaAPI
from finnomena_api.keys import keys
import warnings
import yfinance as yf
import subprocess
import time

api = finnomenaAPI()

df_search_list = pd.read_csv("./csv/input.csv")
code_name_column= df_search_list['code-name'].tolist()
print(1)
search_list = []
for i in range(len(code_name_column)):
    list = code_name_column[i].strip()
    search_list.append(list)
print('we have {0} string for {1}'.format(len(search_list[0]),search_list[0])) #
print('we have {0} name-code to search'.format(len(search_list))) 
df_search_list = pd.read_csv("./csv/input.csv")
code_name_column= df_search_list['y-code'].tolist()


search_list_fed = []
for i in range(len(code_name_column)):
    list = code_name_column[i].strip()
    search_list_fed.append(list)
print('we have {0} string for {1}'.format(len(search_list_fed[0]),search_list_fed[0])) 
print('we have {0} name-code to search'.format(len(search_list_fed)))




missing = []
df_funds_info = pd.DataFrame()
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    for i in range(len(search_list)):
        try:
            run_index = search_list[i]
            dict_info = api.get_fund_info(run_index)
            df_add = pd.DataFrame(data=[dict_info])
            df_funds_info = df_funds_info.append(df_add,ignore_index=True)
        except:
            try:
                print('cannot find: ', run_index)
                missing.append(run_index)
                dict_info =dict.fromkeys(dict_info, 'Error')
                dict_info['security_name'] = str(run_index)
                df_add = pd.DataFrame(data=[dict_info])
                df_funds_info = pd.concat([df_funds_info, df_add], ignore_index=True)
            except:
                print('Caught Error')
        print('processing ', str(run_index))
print('finish processing')
print('missing funds is ', missing)
print('missing funds SUM', len(missing))
print('total number of funds = ',len(df_funds_info))
print('current final reviewed nav_date: ', df_funds_info['nav_date'].max())

from bs4 import BeautifulSoup
fund_types_list =[]
for i in search_list:
    try:
        url = "https://www.finnomena.com/fund/" + str(i)
        page = requests.get(url)
        soup = BeautifulSoup(page.content, 'html.parser')
        soup = soup.find_all('div', class_='detail-row')
        fund_type_add = soup[2].find('div', class_='right').text
        fund_types_list.append(fund_type_add)
    except:
        pass
    print('extracting fund_types:', str(i))
print('finish extraction')
print('total number extract fund_types is: ', len(fund_types_list))

fund_types_df = pd.DataFrame(fund_types_list)
fund_types_df = fund_types_df.rename(columns = {0: 'fund_type'})

df_funds_info = pd.concat([df_funds_info, fund_types_df], axis = 1)


y_fund_df = pd.DataFrame({'y_code':[],'y_nav_date':[],'y_nav':[],'y_link':[]})
y_main_link = 'https://finance.yahoo.com/quote/'
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    for i in search_list_fed:
        try:
            fund = yf.Ticker(str(i))
            fund_his = fund.history(period="max", interval="1d")
            fund_add = fund_his.round(2).sort_index(ascending=False)
            fund_add.reset_index(inplace=True)
            fund_add = fund_add[['Date','Close']]
            fund_add['Date'] = fund_add['Date'].dt.date
            y_code = str(i)
            y_nav = fund_add['Close'].iloc[0]
            y_nav_date = fund_add['Date'].iloc[0]
            y_link = y_main_link + str(i)
            print('y-code: {0} nav: {1} & nav_date: {2}'.format(str(i),y_nav, y_nav_date))
            y_fund_df = y_fund_df.append({'y_code':y_code,'y_nav_date':y_nav_date,'y_nav':y_nav,'y_link':y_link}, ignore_index = True)
        except:
            try:
                y_code = str(i)
                y_nav = ('none')
                y_nav_date = ('none')
                y_link = ('none')
                print('THIS IS NOT FOUND y-code: {0} nav: {1} & nav_date: {2}'.format(str(i),y_nav, y_nav_date))
                y_fund_df = y_fund_df.append({'y_code':y_code,'y_nav_date':y_nav_date,'y_nav':y_nav,'y_link':y_link}, ignore_index = True)
            except:
                print('ERROR')
                break

print('Finish extract nav and nav_date from yahoo site')

with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    df_funds_tracking = df_funds_info[['security_name','nav_date','current_price','fund_type','feeder_fund']]
    df_funds_tracking['mainlink'] = "https://www.finnomena.com/fund/"
    df_funds_tracking['weblink'] = df_funds_tracking['mainlink']+ df_funds_tracking['security_name']
    df_funds_tracking = df_funds_tracking[['security_name','nav_date','current_price','fund_type','feeder_fund','weblink']]
    df_funds_tracking = df_funds_tracking.rename(columns = {'current_price':'nav'})
    df_funds_tracking[['y_code','y_nav','y_nav_date','y_link']] = y_fund_df[['y_code','y_nav','y_nav_date','y_link']]

print('export file: df_funds_tracking.csv')
df_funds_tracking.to_csv('df_funds_tracking.csv')

print('finish extracting nav')

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

to_add_df =pd.read_csv('df_funds_tracking.csv')
to_add_df = to_add_df[['security_name','nav_date','nav']]
print(to_add_df.loc[31,'security_name'])
print(to_add_df.iloc[2,0])


file_name = 'C:/Users/kpornpanom/OneDrive/Desktop/0_MyMoneyManagement(new).xlsx'
wb = load_workbook(file_name)
row_num = pd.read_csv('df_funds_tracking.csv')

row_num = len(row_num)
print('data type: ',type(row_num))
print('data len: ', row_num)

number_of_columns = to_add_df.shape[1]
print('number of columns: ', number_of_columns)

ws = wb['NAV Data']
print('Work book = ', ws)


col= 2 
col_offset = col 
last_col = number_of_columns+2
xl_col = []
while col <= last_col:
    add = get_column_letter(col)
    xl_col.append(add)
    col+=1
col= 2 
xl_col




start_row = 2
row_offset = start_row
for row in range(start_row,row_num):
    for colxl,colpy in zip(xl_col,range(col,last_col)):
        ws[colxl + str(row)].value = to_add_df.iloc[row-row_offset,colpy-col_offset]


wb.save(file_name)

import time
time.sleep(2)

import subprocess
subprocess.Popen(['start', file_name], shell=True)

print('FINISH EXPORTING NAV TO EXCEL FILE')

###############################################################################HISTORY PRICE PART
to_drop = []
for i in range(len(df_search_list)):
    if df_search_list.iloc[i,1] == 'none' or df_search_list.iloc[i,1] == 'to add':
        to_drop.append(i)
df_search_list = df_search_list.drop(df_search_list.index[to_drop])
search_list = df_search_list.iloc[:,0].tolist()
search_list_fed = df_search_list.iloc[:,1].tolist()

for i in range(len(search_list)):
    search_list[i] = search_list[i].strip()
print('we have {0} string for {1}'.format(len(search_list[0]),search_list[0])) 
print('we have {0} name-code to search'.format(len(search_list)))

for i in range(len(search_list_fed)):
    search_list_fed[i] = search_list_fed[i].strip()
print('we have {0} string for {1}'.format(len(search_list_fed[0]),search_list_fed[0])) 
print('we have {0} name-code to search'.format(len(search_list_fed)))

df_funds_price = pd.DataFrame()

missing =[]
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    for i in range(len(search_list)):
        try:
            run_index = search_list[i]
            fund_f = api.get_fund_price(run_index)
            df = fund_f.round(2).sort_index(ascending=False)

            df.set_index("date", inplace=True)

            df.index = pd.to_datetime(df.index)

            df_weekly = df.resample("W-MON").agg({"price": ["first", "max", "min", "last"]})

            df_weekly.columns = ["Open", "High", "Low", "Close Feeder"]
            
            df_weekly = df_weekly.sort_index(ascending=False)
            df_weekly[['Close Fed','Code', 'Security Filter']]=['NaN',run_index,run_index]
            df_weekly['Date']=df_weekly.index
            df_weekly['Date'] = df_weekly['Date'].dt.date
            df_weekly['Year'] = pd.to_datetime(df_weekly['Date']).dt.year
            df_weekly['Categories'] = 'Local'
            df_funds_price = pd.concat([df_funds_price, df_weekly], ignore_index=True)
            print('processed:{0} {1} '.format(i,run_index))
        except:
            try:
                print('Cannot find: ', run_index)
                missing.append(run_index)
                df_weekly =dict.fromkeys(df_weekly, 'Error')
                df_weekly['code'] = str(run_index)
                df_weekly = pd.DataFrame(data=[df_weekly])
                df_funds_price = pd.concat([df_funds_price, df_weekly], ignore_index=True)
            except:
                print('Caught Error')
print('missing funds is ', missing)
print('missing funds SUM', len(missing))
print('total number of funds = ',len(df_funds_price))



df_funds_price_f = pd.DataFrame()
missing =[]
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    for i in range(len(search_list_fed)):
        try:
            run_index = search_list_fed[i]
            run_index_filter = search_list[i]
            fund_y = yf.Ticker(run_index)
            df = fund_y.history(period="max", interval="1d")
            if len(df) == 0:
                print('Cannot find: ', run_index)
                pass
            else:
                df = df.round(2).sort_index(ascending=False)
                df.reset_index(inplace=True)
                df.set_index('Date', inplace=True)
                df.index = pd.to_datetime(df.index)
                df_weekly = df.resample('W-MON').agg({'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last'})
                df_weekly = df_weekly.rename(columns = {'Close':'Close Fed'})
                df_weekly[['Close Feeder','Code', 'Security Filter','Categories']]=['NaN',run_index,run_index_filter,'Global']
                df_weekly['Date'] = df_weekly.index
                df_weekly['Date'] = df_weekly['Date'].dt.date
                df_weekly['Year'] = pd.to_datetime(df_weekly['Date']).dt.year
                df_weekly = df_weekly[['Open','High','Low','Close Feeder','Close Fed','Code','Security Filter','Date','Categories','Year']]
                df_weekly = df_weekly.round(2).sort_index(ascending=False)
                df_funds_price_f = pd.concat([df_funds_price_f, df_weekly], ignore_index=True)
                print('processed:{0} {1} '.format(i,run_index))
        except:
            print('Caught Error')
            break
print('missing funds is ', missing)
print('missing funds SUM', len(missing))
print('total number of funds = ',len(df_funds_price_f))

df_funds_price_union = pd.concat([df_funds_price, df_funds_price_f], ignore_index=True)
df_funds_price_union.dropna(inplace=True)

print(df_funds_price_union.isna().sum())
print(df_funds_price_union.dtypes)
df_funds_price_union['Open'] = df_funds_price_union['Open'].astype(float)
df_funds_price_union['High'] = df_funds_price_union['High'].astype(float)
df_funds_price_union['Low'] = df_funds_price_union['Low'].astype(float)
df_funds_price_union['Close Feeder'] = df_funds_price_union['Close Feeder'].astype(float)
df_funds_price_union['Close Fed'] = df_funds_price_union['Close Fed'].astype(float)
print(df_funds_price_union.dtypes)
print(len(df_funds_price_union))
df_funds_price_union.columns

import pandas as pd
import pygsheets
GSHEET_NAME = 'makeone'

df = df_funds_price_union

creds = 'pkong-credential.json'
api = pygsheets.authorize(service_file=creds)
wb = api.open(GSHEET_NAME)
sheet = wb.worksheet_by_title(f'Sheet1')
sheet.clear(start='A1', end=None, fields='*')
sheet.set_dataframe(df, (1,1))






